import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


def insert_images_to_excel(folder_path, output_excel_path):
    """
    将指定文件夹中的所有图片插入到一个新的 Excel 文件中，每张图片占一行。

    :param folder_path: 图片文件夹路径（字符串或 Path 对象）
    :param output_excel_path: 输出的 Excel 文件路径（如 'output.xlsx'）
    """
    folder = Path(folder_path)
    if not folder.exists() or not folder.is_dir():
        raise ValueError(f"文件夹不存在或无效: {folder_path}")

    # 支持的图片扩展名（小写）
    image_extensions = {'.png', '.jpg', '.jpeg', '.bmp', '.gif'}

    # 获取所有图片文件
    image_files = [
        f for f in folder.iterdir()
        if f.is_file() and f.suffix.lower() in image_extensions
    ]

    if not image_files:
        print("⚠️ 警告：文件夹中未找到支持的图片文件。")
        return

    # 创建新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "图片汇总"

    # 设置标题
    ws.cell(row=1, column=1, value="文件名")
    ws.cell(row=1, column=2, value="图片")

    # 从第2行开始插入
    start_row = 2
    for idx, img_path in enumerate(sorted(image_files)):
        row = start_row + idx

        # 写入文件名
        ws.cell(row=row, column=1, value=img_path.name)

        try:
            # 加载图片
            img = Image(str(img_path))

            # 可选：设置图片缩放（避免过大）
            # openpyxl 不直接支持 resize by pixel, 但可通过 anchor 和比例调整
            # 这里我们限制最大高度为 150 像素（Excel 行高单位 ≈ 0.75 像素/单位）
            max_height_px = 300
            if img.height > max_height_px:
                scale = max_height_px / img.height
                img.width = int(img.width * scale)
                img.height = max_height_px


            # 插入图片到 B 列（第2列）
            ws.add_image(img, f"B{row}")

            # 调整行高（Excel 行高单位 ≈ 0.75 像素，所以 150px ≈ 200 单位）
            ws.row_dimensions[row].height = max(100, int(img.height * 0.8))  # 粗略估算

        except Exception as e:
            print(f"❌ 无法插入图片 {img_path.name}: {e}")
            ws.cell(row=row, column=2, value="【插入失败】")

    # 调整列宽
    ws.column_dimensions['A'].width = 30  # 文件名列
    ws.column_dimensions['B'].width = 60  # 图片列（Excel 列宽单位非像素，仅视觉调整）

    # 保存 Excel
    wb.save(output_excel_path)
    print(f"✅ 成功生成 Excel 文件：{output_excel_path}")
    print(f"📊 共插入 {len(image_files)} 张图片。")


# ===== 使用示例 =====
if __name__ == "__main__":
    # 请修改为你自己的图片文件夹路径
    folder = r"C:\Users\canway\PycharmProjects\ResumeAce\src\screenshot"

    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"{current_date}_jobs-岗位截图.xlsx"

    try:
        insert_images_to_excel(folder, output_file)
    except Exception as e:
        print(f"❌ 发生错误: {e}")